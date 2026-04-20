"""
Step 8: Final Report Generation
"""
from __future__ import annotations

import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from steps.colors import (
    Colors, color, red, green, yellow, blue, magenta, cyan, white,
    bright_red, bright_green, bright_yellow, bright_blue, bright_magenta, bright_cyan
)

# CSV output directory for expert dialogues
EXPERT_DIALOGUES_DIR = "expert_dialogues"


def print_step_header():
    """Print step 8 header."""
    print()
    print("-" * 60)
    print("  Step 8/8: Generate Final Report")
    print("-" * 60)
    print()


def load_json(filepath: Path) -> dict:
    """Load JSON file."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"  [Warning] Failed to read file {filepath}: {e}")
        return {}


def save_json(filepath: Path, data: dict):
    """Save JSON file."""
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def read_markdown(filepath: Path) -> str:
    """Read markdown file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return f.read()


def read_text(filepath: Path) -> str:
    """Read text file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return f.read()


def save_tables_as_csv(csv_path: Path, sheets_data: List[Dict]) -> bool:
    """
    Save analysis tables as CSV files (one file per sheet/table).

    Args:
        csv_path: Base path for CSV file (e.g., run_dir / "analysis_tables.csv")
        sheets_data: List of dicts with keys: name, headers, rows

    Returns:
        True if successful
    """
    try:
        # For multiple sheets, create separate CSV files with numbered prefixes
        for i, sheet in enumerate(sheets_data, 1):
            sheet_name = sheet.get("name", f"Sheet{i}")
            headers = sheet.get("headers", [])
            rows = sheet.get("rows", [])

            # Sanitize sheet name for filename
            safe_name = sheet_name.replace("/", "_").replace("\\", "_")

            if len(sheets_data) > 1:
                # Multiple sheets: create separate files like analysis_tables_1_criteria.csv
                sheet_csv_path = csv_path.parent / f"analysis_tables_{i}_{safe_name}.csv"
            else:
                # Single sheet: use the csv_path directly
                sheet_csv_path = csv_path

            with open(sheet_csv_path, 'w', encoding='utf-8', newline='') as f:
                # Write BOM for Excel compatibility
                f.write('\ufeff')

                # Write headers
                if headers:
                    f.write(','.join(f'"{h}"' for h in headers) + '\n')

                # Write data rows
                for row in rows:
                    row_str = ','.join(
                        f'"{cell}"' if cell is not None else '""'
                        for cell in row
                    )
                    f.write(row_str + '\n')

            print(f"  [已生成] {sheet_csv_path}")

        return True
    except Exception as e:
        print(f"    [警告] CSV保存失败: {e}")
        return False


def _make_matrix_sheet(
    wb, sheet_name: str,
    row_labels: List[str],
    col_labels: List[str],
    combined: List[List[float]],  # [row][col] = geometric mean Saaty value
    tab_color: str,
) -> None:
    """
    Create a full n×n pairwise comparison matrix sheet.

    Each cell displays only the combined Saaty value (geometric mean).
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    SHEET_HEADER_FILL = PatternFill(start_color='4472C4', fill_type='solid')
    ROW_LABEL_FILL = PatternFill(start_color='D9E1F2', fill_type='solid')
    WHITE_FILL = PatternFill(start_color='FFFFFF', fill_type='solid')
    DIAGONAL_FILL = PatternFill(start_color='E2EFDA', fill_type='solid')

    HEADER_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
    SAATY_FONT = Font(name='Microsoft YaHei', bold=True, size=11, color='000000')
    ROW_LABEL_FONT = Font(name='Microsoft YaHei', bold=True, size=9, color='000000')

    THIN = Side(style='thin')
    THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color

    n_rows = len(row_labels)
    n_cols = len(col_labels)

    # Column widths
    ws.column_dimensions['A'].width = 22
    for c in range(n_cols):
        ws.column_dimensions[get_column_letter(c + 2)].width = 14

    # Row 1: header
    ws.row_dimensions[1].height = 32
    corner = ws.cell(row=1, column=1)
    corner.value = "Pairwise Comparison"
    corner.font = HEADER_FONT
    corner.fill = SHEET_HEADER_FILL
    corner.alignment = CENTER
    corner.border = THIN_BORDER

    for j, col_label in enumerate(col_labels, 2):
        cell = ws.cell(row=1, column=j)
        cell.value = col_label
        cell.font = HEADER_FONT
        cell.fill = SHEET_HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.freeze_panes = 'B2'

    # Data rows
    for i, row_label in enumerate(row_labels, 2):
        ws.row_dimensions[i].height = 28
        rl = ws.cell(row=i, column=1)
        rl.value = row_label
        rl.font = ROW_LABEL_FONT
        rl.fill = ROW_LABEL_FILL
        rl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        rl.border = THIN_BORDER

        for j in range(n_cols):
            cell = ws.cell(row=i, column=j + 2)
            saaty_val = combined[i - 2][j] if (i - 2 < len(combined) and j < len(combined[i - 2])) else 1.0

            if i - 2 == j:
                cell.value = 1
                cell.font = Font(name='Microsoft YaHei', size=11, color='595959', italic=True)
                cell.fill = DIAGONAL_FILL
            else:
                cell.value = round(saaty_val, 2)
                cell.font = SAATY_FONT

            cell.alignment = CENTER
            cell.border = THIN_BORDER

    wb.active = ws


def save_tables_as_xlsx(xlsx_path: Path, sheets_data: List[Dict]) -> bool:
    """
    Save all tables as a single multi-sheet XLSX file (1000minds standard format).

    Supports two sheet types:
      - "matrix": full n×n pairwise comparison matrix with combined Saaty value
      - "list": tabular format (default)
    """
    try:
        import openpyxl
        from openpyxl.styles import (Font, Alignment, PatternFill, Border, Side,
                                     numbers, GradientFill)
        from openpyxl.utils import get_column_letter
        from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

        wb = openpyxl.Workbook()
        if wb.active:
            wb.remove(wb.active)

        # ── Style Definitions (minimax-xlsx financial standard) ──────────────
        HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4',
                                  fill_type='solid')
        SUBHEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2',
                                    fill_type='solid')
        ALT_ROW_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2',
                                   fill_type='solid')
        WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF',
                                 fill_type='solid')

        HEADER_FONT = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
        SUBHEADER_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='000000')
        BLUE_FONT = Font(name='Microsoft YaHei', size=10, color='0000FF')  # input
        BLACK_FONT = Font(name='Microsoft YaHei', size=10, color='000000')  # formula
        GREEN_FONT = Font(name='Microsoft YaHei', size=10, color='008000')  # cross-sheet
        LABEL_FONT = Font(name='Microsoft YaHei', size=10, color='000000')

        THIN = Side(style='thin')
        MEDIUM = Side(style='medium')
        THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        MEDIUM_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
        BOTTOM_MEDIUM = Border(left=THIN, right=THIN, top=THIN, bottom=MEDIUM)

        CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
        RIGHT = Alignment(horizontal='right', vertical='center')

        FMT_PERCENT = '0.00%'
        FMT_PERCENT_1 = '0.0%'
        FMT_NUMBER_4 = '0.0000'
        FMT_NUMBER_2 = '0.00'
        FMT_NUMBER_1 = '0.0'

        def style_cell(cell, font=None, fill=None, border=None, alignment=None,
                       number_format=None):
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment
            if number_format:
                cell.number_format = number_format

        # ── Build Sheets ────────────────────────────────────────────────────
        tab_colors = ['4472C4', '70AD47', 'ED7D31', 'A9D18E',
                      'FFC000', '5B9BD5', 'C55A11', '7030A0',
                      '00B0F0', '00B050']

        for i, sheet in enumerate(sheets_data, 1):
            # Only matrix sheets go to XLSX; list/CSV sheets are already in separate CSV files
            if sheet.get("type") != "matrix":
                continue

            sheet_name = sheet.get("name", f"Sheet{i}")
            # Excel sheet name max 31 chars
            sheet_name = sheet.get("xlsx_name", sheet.get("name", f"Sheet{i}"))
            sheet_name = sheet_name[:31].replace('/', '-').replace('\\', '-')
            headers = sheet.get("headers", [])
            rows = sheet.get("rows", [])

            # Handle matrix sheets specially (full n×n pairwise comparison)
            if sheet.get("type") == "matrix":
                row_labels = sheet.get("row_labels", [])
                col_labels = sheet.get("col_labels", [])
                combined = sheet.get("combined", [])  # [row][col] geo mean
                tab_color = tab_colors[(i - 1) % len(tab_colors)]
                _make_matrix_sheet(wb, sheet_name, row_labels, col_labels,
                                   combined, tab_color)
                continue

            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_properties.tabColor = tab_colors[(i - 1) % len(tab_colors)]

            ws = wb.create_sheet(title=sheet_name)

            # Detect column types from data
            has_numbers = any(
                isinstance(cell, (int, float))
                for row in rows for cell in row
                if cell not in (None, '', '-')
            )

            # ── Header Row ──────────────────────────────────────────────
            ws.row_dimensions[1].height = 30
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                style_cell(cell,
                           font=HEADER_FONT,
                           fill=HEADER_FILL,
                           border=THIN_BORDER,
                           alignment=CENTER)
                # Merge if consecutive headers share a prefix pattern
            ws.freeze_panes = 'A2'

            # ── Data Rows ────────────────────────────────────────────────
            for row_idx, row in enumerate(rows, 2):
                ws.row_dimensions[row_idx].height = 18
                fill = WHITE_FILL if row_idx % 2 == 0 else ALT_ROW_FILL
                for col_idx, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    style_cell(cell, fill=fill, border=THIN_BORDER)

                    if cell_value is None or cell_value == '':
                        cell.value = None
                        style_cell(cell, font=LABEL_FONT, alignment=CENTER)
                    elif cell_value == '-':
                        cell.value = '-'
                        style_cell(cell, font=LABEL_FONT, alignment=CENTER)
                    elif isinstance(cell_value, (int, float)):
                        cell.value = cell_value
                        # Determine if this looks like a percentage or weight
                        if 0 <= cell_value <= 1:
                            style_cell(cell, font=BLACK_FONT, alignment=RIGHT,
                                       number_format=FMT_NUMBER_4)
                        elif cell_value > 1:
                            style_cell(cell, font=BLACK_FONT, alignment=RIGHT,
                                       number_format=FMT_NUMBER_2)
                        else:
                            style_cell(cell, font=BLACK_FONT, alignment=RIGHT)
                    else:
                        cell.value = str(cell_value)
                        style_cell(cell, font=LABEL_FONT, alignment=LEFT)

            # ── Column Widths ────────────────────────────────────────────
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 10
                for row in rows:
                    if col_idx <= len(row) and row[col_idx - 1] is not None:
                        max_len = max(max_len, len(str(row[col_idx - 1])))
                width = min(max(max_len + 2, 8), 35)
                ws.column_dimensions[col_letter].width = width

            ws.sheet_properties.tabColor = tab_colors[(i - 1) % len(tab_colors)]

        wb.save(xlsx_path)
        print(f"  [已生成] {xlsx_path}")
        return True
    except Exception as e:
        print(f"    [警告] XLSX保存失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def _geom_mean(values: List[float]) -> float:
    """Compute geometric mean of a list of positive numbers."""
    import math
    values = [v for v in values if v and v > 0]
    if not values:
        return 1.0
    return math.prod(values) ** (1.0 / len(values))


def _build_full_matrix(
    items: List[dict],
    expert_data: List[dict],
    get_value_fn,
) -> Tuple[List[List[float]], List[List[List[float]]]]:
    """
    Build full n×n Saaty comparison matrix.

    Returns:
        combined: n×n matrix of geometric means
        expert_matrices: list of n×n matrices, one per expert
    """
    import math
    n = len(items)
    combined = [[1.0 for _ in range(n)] for _ in range(n)]
    expert_matrices = []

    for ed in expert_data:
        mat = [[1.0 for _ in range(n)] for _ in range(n)]
        for i in range(n):
            for j in range(n):
                if i != j:
                    mat[i][j] = get_value_fn(ed, i, j)
        # Normalize rows to get local priorities (for display/sorting)
        row_sums = [sum(mat[i][j] for j in range(n)) for i in range(n)]
        for i in range(n):
            for j in range(n):
                mat[i][j] = mat[i][j] / row_sums[i] if row_sums[i] > 0 else 0
        expert_matrices.append(mat)

    # Build geometric mean matrix
    for i in range(n):
        for j in range(n):
            if i != j:
                vals = [get_value_fn(ed, i, j) for ed in expert_data]
                combined[i][j] = _geom_mean(vals)

    return combined, expert_matrices


def generate_analysis_csv(state: dict, run_dir: Path) -> Path:
    """
    Generate analysis tables in 1000minds AHP standard format.

    Sheet structure (pairwise_comparisons format):
      1.  准则层两两比较矩阵 (full n×n criteria matrix with expert Saaty breakdown)
      2.  几何平均优先向量 (geometric mean priority vector)
      3.  准则层权重及一致性 (criteria weights + CI/CR)
      4.  方案层-科研创新质量 (alternatives Saaty matrix under C01)
      5.  方案层-团队协同与支撑 (alternatives Saaty matrix under C02)
      6.  方案层-成果转化与产业贡献 (alternatives Saaty matrix under C03)
      7.  方案层-社会服务与战略契合 (alternatives Saaty matrix under C04)
      8.  方案层-人才培养与成果推广 (alternatives Saaty matrix under C05)
      9.  方案层局部权重(专家明细) (expert disaggregation of local weights)
      10. 综合权重排名 (all alternatives with local weights + combined)
    """
    import math

    # Load all data
    hierarchy = load_json(run_dir / "ahp_hierarchy.json") if (run_dir / "ahp_hierarchy.json").exists() else {}
    ahp_results = load_json(run_dir / "ahp_results.json") if (run_dir / "ahp_results.json").exists() else {}
    criteria_comparisons = load_json(run_dir / "criteria_comparisons.json") if (run_dir / "criteria_comparisons.json").exists() else {}
    alternative_scores = (load_json(run_dir / "alternative_scores.json") if (run_dir / "alternative_scores.json").exists() else {}) or {}

    criteria = hierarchy.get("criteria_layer", [])
    alternatives = hierarchy.get("alternative_layer", [])
    criteria_weights_list = ahp_results.get("criteria", [])
    criteria_map = {c["id"]: c["name"] for c in criteria}
    n = len(criteria)
    comparisons = criteria_comparisons.get("comparisons", {})
    expert_comparisons = criteria_comparisons.get("expert_comparisons", [])
    fmt = alternative_scores.get("format", "")

    # ─────────────────────────────────────────────────────────────────
    # Helper: build expert lookup for criteria comparisons
    # expert_comparisons[i]["comparisons"] = [{pair: [id_i, id_j], value}]
    # ─────────────────────────────────────────────────────────────────
    crit_ids = [c["id"] for c in criteria]

    def get_crit_saaty(ed: dict, i: int, j: int) -> float:
        """Get Saaty comparison value from expert data for criteria i,j."""
        cid_i = crit_ids[i]
        cid_j = crit_ids[j]
        for comp in ed.get("comparisons", []):
            p = comp.get("pair", [])
            if len(p) == 2:
                if (p[0] == cid_i and p[1] == cid_j):
                    return comp.get("value", 1.0)
                if (p[0] == cid_j and p[1] == cid_i):
                    return 1.0 / comp.get("value", 1.0) if comp.get("value", 0) != 0 else 1.0
        return 1.0

    all_sheets = []

    # ─────────────────────────────────────────────────────────────────
    # Sheet 1: 准则层两两比较矩阵 (full n×n matrix + pairwise CSV)
    # ─────────────────────────────────────────────────────────────────
    expert_names = [ec.get("expert_name", f"专家{ec.get('expert_id','?')}") for ec in expert_comparisons]
    n_exp = len(expert_comparisons)

    # Build full n×n matrix: [row][col][expert_vals]
    crit_matrix = [[[1.0] * n_exp for _ in range(n)] for _ in range(n)]
    crit_combined = [[1.0] * n for _ in range(n)]

    for i in range(n):
        for j in range(n):
            if i != j:
                vals = [get_crit_saaty(ec, i, j) for ec in expert_comparisons]
                for k, v in enumerate(vals):
                    crit_matrix[i][j][k] = v
                crit_combined[i][j] = _geom_mean([v for v in vals if v and v > 0])

    # XLSX matrix format
    sheet1_matrix = {
        "type": "matrix",
        "name": "表1-准则层两两比较矩阵",
        "row_labels": [c["name"] for c in criteria],
        "col_labels": [c["name"] for c in criteria],
        "expert_names": expert_names,
        "matrix_data": crit_matrix,   # [row][col][expert_vals]
        "combined": crit_combined,   # [row][col] geo mean
    }

    # CSV pairwise list format (flattened upper triangle)
    sheet1_csv_headers = ["准则A", "准则B"] + expert_names + ["Saaty值(几何平均)"]
    sheet1_csv_rows = []
    for i in range(n):
        for j in range(i + 1, n):
            c1_name, c2_name = criteria[i]["name"], criteria[j]["name"]
            expert_vals = [round(crit_matrix[i][j][k], 2) for k in range(n_exp)]
            geo_val = crit_combined[i][j]
            sheet1_csv_rows.append([c1_name, c2_name] + expert_vals + [round(geo_val, 4)])

    sheet1_csv = {
        "name": "表1-准则层两两比较矩阵",
        "headers": sheet1_csv_headers,
        "rows": sheet1_csv_rows
    }

    all_sheets.append(sheet1_matrix)
    all_sheets.append(sheet1_csv)

    # ─────────────────────────────────────────────────────────────────
    # Sheet 2: 几何平均优先向量
    # ─────────────────────────────────────────────────────────────────
    # Compute geo means from the combined matrix rows
    sheet2_geo_means = []
    for i in range(n):
        product = 1.0
        for j in range(n):
            product *= (comparisons.get(f"{min(i,j)},{max(i,j)}", 1.0) if i != j else 1.0)
        sheet2_geo_means.append(product ** (1.0 / n) if product > 0 else 1e-10)
    total_geo = sum(sheet2_geo_means)
    norm_weights = [w / total_geo for w in sheet2_geo_means]

    sheet2_headers = ["编号", "准则名称", "几何平均值", "归一化权重"]
    sheet2_rows = []
    for i, c in enumerate(criteria):
        sheet2_rows.append([
            c["id"], c["name"],
            round(sheet2_geo_means[i], 6),
            round(norm_weights[i], 4)
        ])

    sheet2 = {
        "name": "表2-几何平均优先向量",
        "headers": sheet2_headers,
        "rows": sheet2_rows
    }
    all_sheets.append(sheet2)

    # ─────────────────────────────────────────────────────────────────
    # Sheet 3: 准则层权重及一致性
    # ─────────────────────────────────────────────────────────────────
    crit_cons = ahp_results.get("criteria_consistency", {})
    sheet3_headers = ["编号", "准则名称", "权重", "λmax", "CI", "CR", "一致性"]
    sheet3_rows = []
    for c in criteria:
        cw = next((cr for cr in criteria_weights_list if cr["id"] == c["id"]), {})
        cr_val = cw.get("cr", 0)
        lambda_val = cw.get("lambda_max", 0)
        ci_val = cw.get("ci", 0)
        passed = "通过" if cw.get("consistency_passed", False) else "未通过"
        sheet3_rows.append([
            c["id"], c["name"],
            round(cw.get("weight", 0), 4),
            round(lambda_val, 4),
            round(ci_val, 4),
            round(cr_val, 4),
            passed
        ])
    # Summary row
    sheet3_rows.append([
        "", "汇总",
        round(crit_cons.get("lambda_max", 0), 4), "",
        round(crit_cons.get("ci", 0), 4),
        round(crit_cons.get("cr", 0), 4),
        "通过" if crit_cons.get("passed") else "未通过"
    ])

    sheet3 = {
        "name": "表3-准则层权重及一致性",
        "headers": sheet3_headers,
        "rows": sheet3_rows
    }
    all_sheets.append(sheet3)

    # ─────────────────────────────────────────────────────────────────
    # Sheets 4-8: Per-criterion alternatives Saaty comparison matrices
    # Each sheet: full n×n matrix in XLSX, pairwise list in CSV
    # ─────────────────────────────────────────────────────────────────
    local_weights = alternative_scores.get("local_weights", {})
    alt_by_crit = alternative_scores.get("alternatives_by_criteria", {})
    expert_details = alternative_scores.get("expert_details", [])

    def get_alt_saaty(ed: dict, crit_id: str, i: int, j: int) -> float:
        mat = ed.get("criterion_matrices", {}).get(crit_id, [])
        if mat and i < len(mat) and j < len(mat[i]):
            return mat[i][j]
        return 1.0

    sheet_idx = 4
    for crit in criteria:
        cid = crit["id"]
        alts = alt_by_crit.get(cid, [])
        if not alts:
            continue
        m = len(alts)
        alt_names = [a["name"] for a in alts]

        # Build full m×m matrix: [row][col][expert_vals]
        alt_matrix = [[[1.0] * n_exp for _ in range(m)] for _ in range(m)]
        alt_combined = [[1.0] * m for _ in range(m)]

        for i in range(m):
            for j in range(m):
                if i != j:
                    vals = [get_alt_saaty(ed, cid, i, j) for ed in expert_details]
                    for k, v in enumerate(vals):
                        alt_matrix[i][j][k] = v
                    alt_combined[i][j] = _geom_mean([v for v in vals if v and v > 0])

        # XLSX matrix format
        sheet_matrix = {
            "type": "matrix",
            "name": f"表{sheet_idx}-{crit['name'][:20]}",
            "row_labels": alt_names,
            "col_labels": alt_names,
            "expert_names": expert_names,
            "matrix_data": alt_matrix,
            "combined": alt_combined,
        }

        # CSV pairwise list
        csv_headers = ["方案A", "方案B"] + expert_names + ["Saaty值(几何平均)"]
        csv_rows = []
        for i in range(m):
            for j in range(i + 1, m):
                a1_name, a2_name = alt_names[i], alt_names[j]
                exp_vals = [round(alt_matrix[i][j][k], 2) for k in range(n_exp)]
                geo_val = alt_combined[i][j]
                csv_rows.append([a1_name, a2_name] + exp_vals + [round(geo_val, 4)])

        sheet_csv = {
            "name": f"表{sheet_idx}-{crit['name'][:20]}",
            "headers": csv_headers,
            "rows": csv_rows
        }

        all_sheets.append(sheet_matrix)
        all_sheets.append(sheet_csv)
        sheet_idx += 1

    # ─────────────────────────────────────────────────────────────────
    # Sheet 9: 方案层局部权重(专家明细)
    # ─────────────────────────────────────────────────────────────────
    sheet9_headers = ["准则", "编号", "方案名称"] + expert_names + ["聚合权重(几何平均)"]
    sheet9_rows = []
    for crit in criteria:
        cid = crit["id"]
        alts = alt_by_crit.get(cid, [])
        if not alts:
            continue
        for alt in alts:
            aid = alt["id"]
            aname = alt["name"]
            row = [cid, aid, aname]
            for ed in expert_details:
                ew = ed.get("local_weights", {}).get(cid, {}).get(aid, 0.0)
                row.append(round(ew, 4))
            row.append(round(local_weights.get(cid, {}).get(aid, 0.0), 4))
            sheet9_rows.append(row)

    sheet9 = {
        "name": f"表{sheet_idx}-方案层局部权重(专家明细)",
        "headers": sheet9_headers,
        "rows": sheet9_rows
    }
    all_sheets.append(sheet9)
    sheet_idx += 1

    # ─────────────────────────────────────────────────────────────────
    # Sheet 10: 综合权重排名（1000minds格式：各准则局部权重展开）
    # ─────────────────────────────────────────────────────────────────
    ranking = ahp_results.get("ranking", [])
    rank_map = {r["id"]: r["rank"] for r in ranking}
    sorted_by_rank = sorted(alternatives, key=lambda x: rank_map.get(x["id"], 999))

    # Build alt_local_weights lookup: alt_id -> {crit_id: local_weight}
    alt_local_weights = {}
    if fmt == "pairwise_comparisons":
        for crit_id, alts_dict in alt_by_crit.items():
            for alt in alts_dict:
                aid = alt["id"]
                if aid not in alt_local_weights:
                    alt_local_weights[aid] = {}
                alt_local_weights[aid][crit_id] = local_weights.get(crit_id, {}).get(aid, 0.0)

    sheet10_headers = ["排名", "编号", "方案名称"] + [c["name"] for c in criteria] + ["综合权重"]
    sheet10_rows = []
    for alt in sorted_by_rank:
        aid = alt["id"]
        aw = next((a for a in ahp_results.get("alternatives", []) if a["id"] == aid), {})
        rank = rank_map.get(aid, "?")
        row = [rank, aid, alt["name"]]
        for c in criteria:
            lw = alt_local_weights.get(aid, {}).get(c["id"], 0.0)
            row.append(round(lw, 4))
        row.append(round(aw.get("combined_weight", 0), 4))
        sheet10_rows.append(row)

    sheet10 = {
        "name": f"表{sheet_idx}-综合权重排名",
        "headers": sheet10_headers,
        "rows": sheet10_rows
    }
    all_sheets.append(sheet10)

    # ─────────────────────────────────────────────────────────────────
    # Create CSV files
    # ─────────────────────────────────────────────────────────────────
    # CSV: only list-type sheets (matrix sheets have their own CSV counterpart)
    csv_sheets = [s for s in all_sheets if s.get("type") != "matrix"]
    csv_path = run_dir / "analysis_tables.csv"
    success = save_tables_as_csv(csv_path, csv_sheets)

    # XLSX: all sheets (including matrix sheets rendered as n×n grids)
    xlsx_path = run_dir / "analysis_tables.xlsx"
    save_tables_as_xlsx(xlsx_path, all_sheets)

    if success:
        return csv_path
    else:
        return None


def generate_interactive_html_report(state: dict, run_dir: Path) -> str:
    """
    Generate an interactive single-file HTML report.
    Following design-taste-frontend principles.
    """
    # Load project from file (not state, which may be empty when called directly)
    project = load_json(run_dir / "project.json") if (run_dir / "project.json").exists() else {}
    project_title = project.get('title', '未知项目') if isinstance(project, dict) else getattr(project, 'title', '未知项目')

    # Load all data from run_dir JSON files
    hierarchy = load_json(run_dir / "ahp_hierarchy.json") if (run_dir / "ahp_hierarchy.json").exists() else {}
    ahp_results = load_json(run_dir / "ahp_results.json") if (run_dir / "ahp_results.json").exists() else {}
    criteria_comparisons = load_json(run_dir / "criteria_comparisons.json") if (run_dir / "criteria_comparisons.json").exists() else {}
    alternative_scores = (load_json(run_dir / "alternative_scores.json") if (run_dir / "alternative_scores.json").exists() else {}) or {}
    convergence_data = load_json(run_dir / "convergence_check.json") if (run_dir / "convergence_check.json").exists() else {}
    sensitivity_data = load_json(run_dir / "sensitivity_analysis.json") if (run_dir / "sensitivity_analysis.json").exists() else {}
    experts_data = load_json(run_dir / "experts.json") if (run_dir / "experts.json").exists() else {}

    # Get expert info from experts.json file
    experts_list = experts_data if isinstance(experts_data, list) else experts_data.get("experts", [])
    expert_list = []
    for exp in experts_list:
        if isinstance(exp, dict):
            expert_list.append({
                "id": exp.get('id', '?'),
                "name": exp.get('name', '未知'),
                "role": exp.get('role', '未知'),
                "org_type": exp.get('org_type', '未知'),
                "expertise": exp.get('expertise', '未知'),
            })
        else:
            expert_list.append({
                "id": getattr(exp, 'id', '?'),
                "name": getattr(exp, 'name', '未知'),
                "role": getattr(exp, 'role', '未知'),
                "org_type": getattr(exp, 'org_type', '未知'),
                "expertise": getattr(exp, 'expertise', '未知'),
            })

    criteria = hierarchy.get("criteria_layer", [])
    alternatives = hierarchy.get("alternative_layer", [])
    ranking = ahp_results.get("ranking", [])
    rank_map = {r["id"]: r["rank"] for r in ranking}
    sorted_alts = sorted(alternatives, key=lambda x: rank_map.get(x["id"], 999))
    criteria_weights = ahp_results.get("criteria", [])
    criteria_map = {c["id"]: c["name"] for c in criteria}

    # Build convergence data using ahp_results CR (not stale convergence_check.json)
    crit_cons = ahp_results.get("criteria_consistency", {})
    conv_with_ahp_cr = dict(convergence_data)
    conv_with_ahp_cr["ahp_consistency"] = {
        "cr": crit_cons.get("cr", 0),
        "lambda_max": crit_cons.get("lambda_max", 0),
        "ci": crit_cons.get("ci", 0),
        "passed": crit_cons.get("passed", False),
    }

    # Build JSON data for JavaScript
    report_data = {
        "project_title": project_title,
        "framework": project.get('framework', '') if isinstance(project, dict) else getattr(project, 'framework', ''),
        "background": project.get('background', '') if isinstance(project, dict) else getattr(project, 'background', ''),
        "purpose": project.get('purpose', '') if isinstance(project, dict) else getattr(project, 'purpose', ''),
        "experts": expert_list,
        "criteria": criteria,
        "criteria_weights": criteria_weights,
        "criteria_map": criteria_map,
        "alternatives": alternatives,
        "sorted_alts": sorted_alts,
        "ranking": ranking,
        "convergence_data": conv_with_ahp_cr,
        "sensitivity_data": sensitivity_data,
    }

    # Serialize data for JS (no indent to avoid JS syntax errors with newlines)
    json_data = json.dumps(report_data, ensure_ascii=False, indent=None, separators=(',', ':'))

    # Generate HTML
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{project_title} - Delphi-AHP 研究报告</title>
    <style>
        :root {{
            --bg-base: #f9fafb;
            --bg-card: #ffffff;
            --border: #e5e7eb;
            --text-primary: #111827;
            --text-secondary: #6b7280;
            --accent: #059669;
            --accent-light: #d1fae5;
            --warning: #f59e0b;
            --danger: #ef4444;
            --radius: 1rem;
        }}

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: "Geist", "Satoshi", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            background: var(--bg-base);
            color: var(--text-primary);
            line-height: 1.6;
            -webkit-font-smoothing: antialiased;
        }}

        .container {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 2rem;
        }}

        /* Header */
        .header {{
            text-align: center;
            padding: 4rem 2rem;
            background: linear-gradient(135deg, #111827 0%, #1f2937 100%);
            color: white;
            border-radius: 0 0 2rem 2rem;
            margin-bottom: 3rem;
        }}

        .header h1 {{
            font-size: 2.5rem;
            font-weight: 700;
            letter-spacing: -0.02em;
            margin-bottom: 0.5rem;
        }}

        .header .subtitle {{
            font-size: 1.1rem;
            opacity: 0.8;
        }}

        .header .meta {{
            margin-top: 1.5rem;
            font-size: 0.875rem;
            opacity: 0.6;
        }}

        /* Navigation Tabs */
        .nav-tabs {{
            display: flex;
            gap: 0.5rem;
            margin-bottom: 2rem;
            flex-wrap: wrap;
        }}

        .nav-tab {{
            padding: 0.75rem 1.5rem;
            border: 1px solid var(--border);
            background: var(--bg-card);
            border-radius: 100px;
            cursor: pointer;
            font-size: 0.875rem;
            font-weight: 500;
            color: var(--text-secondary);
            transition: all 0.2s cubic-bezier(0.16, 1, 0.3, 1);
        }}

        .nav-tab:hover {{
            border-color: var(--accent);
            color: var(--accent);
        }}

        .nav-tab.active {{
            background: var(--accent);
            border-color: var(--accent);
            color: white;
        }}

        /* Cards */
        .card {{
            background: var(--bg-card);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 2rem;
            margin-bottom: 1.5rem;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
        }}

        .card-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 1.5rem;
            padding-bottom: 1rem;
            border-bottom: 1px solid var(--border);
        }}

        .card-title {{
            font-size: 1.25rem;
            font-weight: 600;
            letter-spacing: -0.01em;
        }}

        .badge {{
            padding: 0.25rem 0.75rem;
            border-radius: 100px;
            font-size: 0.75rem;
            font-weight: 600;
        }}

        .badge-success {{
            background: var(--accent-light);
            color: var(--accent);
        }}

        .badge-warning {{
            background: #fef3c7;
            color: var(--warning);
        }}

        .badge-danger {{
            background: #fee2e2;
            color: var(--danger);
        }}

        /* Tables */
        .data-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .data-table th {{
            text-align: left;
            padding: 0.75rem 1rem;
            background: #f3f4f6;
            font-weight: 600;
            font-size: 0.875rem;
            color: var(--text-secondary);
            border-bottom: 2px solid var(--border);
        }}

        .data-table td {{
            padding: 0.75rem 1rem;
            border-bottom: 1px solid var(--border);
        }}

        .data-table tr:hover {{
            background: #f9fafb;
        }}

        /* Grid Layout */
        .grid-2 {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 1.5rem;
        }}

        /* Stat Cards */
        .stat-card {{
            background: var(--bg-card);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 1.5rem;
            text-align: center;
        }}

        .stat-value {{
            font-size: 2.5rem;
            font-weight: 700;
            color: var(--accent);
            letter-spacing: -0.02em;
        }}

        .stat-label {{
            font-size: 0.875rem;
            color: var(--text-secondary);
            margin-top: 0.25rem;
        }}

        /* Progress Bars */
        .progress-bar {{
            height: 8px;
            background: #e5e7eb;
            border-radius: 100px;
            overflow: hidden;
            margin: 0.5rem 0;
        }}

        .progress-fill {{
            height: 100%;
            background: var(--accent);
            border-radius: 100px;
            transition: width 0.6s cubic-bezier(0.16, 1, 0.3, 1);
        }}

        /* Ranking List */
        .ranking-list {{
            list-style: none;
        }}

        .ranking-item {{
            display: flex;
            align-items: center;
            gap: 1rem;
            padding: 1rem;
            border-bottom: 1px solid var(--border);
        }}

        .ranking-item:last-child {{
            border-bottom: none;
        }}

        .rank-badge {{
            width: 2rem;
            height: 2rem;
            display: flex;
            align-items: center;
            justify-content: center;
            border-radius: 50%;
            font-weight: 700;
            font-size: 0.875rem;
        }}

        .rank-1 {{ background: #fef3c7; color: #d97706; }}
        .rank-2 {{ background: #f3f4f6; color: #6b7280; }}
        .rank-3 {{ background: #fed7aa; color: #ea580c; }}
        .rank-default {{ background: #f3f4f6; color: #9ca3af; }}

        .ranking-content {{
            flex: 1;
        }}

        .ranking-name {{
            font-weight: 600;
        }}

        .ranking-meta {{
            font-size: 0.75rem;
            color: var(--text-secondary);
        }}

        .ranking-score {{
            font-weight: 700;
            color: var(--accent);
        }}

        /* Section Content */
        .section-content {{
            display: none;
        }}

        .section-content.active {{
            display: block;
        }}

        /* Criteria Bars */
        .criteria-bar {{
            margin-bottom: 1.5rem;
        }}

        .criteria-label {{
            display: flex;
            justify-content: space-between;
            margin-bottom: 0.5rem;
            font-size: 0.875rem;
        }}

        /* Expert Cards */
        .expert-card {{
            display: flex;
            gap: 1rem;
            padding: 1rem;
            border: 1px solid var(--border);
            border-radius: var(--radius);
            margin-bottom: 1rem;
        }}

        .expert-avatar {{
            width: 3rem;
            height: 3rem;
            border-radius: 50%;
            background: linear-gradient(135deg, #059669, #10b981);
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: 700;
            font-size: 1.25rem;
        }}

        .expert-info {{
            flex: 1;
        }}

        .expert-name {{
            font-weight: 600;
        }}

        .expert-role {{
            font-size: 0.875rem;
            color: var(--text-secondary);
        }}

        /* Footer */
        .footer {{
            text-align: center;
            padding: 3rem 2rem;
            color: var(--text-secondary);
            font-size: 0.875rem;
        }}

        /* Responsive */
        @media (max-width: 768px) {{
            .container {{
                padding: 1rem;
            }}
            .header {{
                padding: 2rem 1rem;
                border-radius: 0 0 1rem 1rem;
            }}
            .header h1 {{
                font-size: 1.75rem;
            }}
            .nav-tabs {{
                justify-content: center;
            }}
            .data-table {{
                font-size: 0.875rem;
            }}
            .data-table th, .data-table td {{
                padding: 0.5rem;
            }}
        }}

        /* Animations */
        @keyframes fadeIn {{
            from {{ opacity: 0; transform: translateY(10px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}

        .card {{
            animation: fadeIn 0.4s cubic-bezier(0.16, 1, 0.3, 1);
        }}

        /* Print */
        @media print {{
            .nav-tabs, .no-print {{ display: none !important; }}
            .card {{ break-inside: avoid; }}
            .header {{ background: #111827 !important; color: white !important; print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>{project_title}</h1>
        <p class="subtitle">Delphi-AHP 研究分析报告</p>
        <p class="meta">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
    </div>

    <div class="container">
        <nav class="nav-tabs no-print">
            <button class="nav-tab active" onclick="showSection('overview')">研究概览</button>
            <button class="nav-tab" onclick="showSection('criteria')">准则层</button>
            <button class="nav-tab" onclick="showSection('alternatives')">方案层</button>
            <button class="nav-tab" onclick="showSection('convergence')">收敛检验</button>
            <button class="nav-tab" onclick="showSection('sensitivity')">敏感性</button>
            <button class="nav-tab" onclick="showSection('experts')">专家团队</button>
        </nav>

        <!-- Overview Section -->
        <div id="overview" class="section-content active">
            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">研究背景</h2>
                </div>
                <p style="color: var(--text-secondary); line-height: 1.8;">{getattr(project, 'background', '') if hasattr(project, 'background') else project.get('background', '未提供')}</p>
            </div>

            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">研究目的</h2>
                </div>
                <p style="color: var(--text-secondary); line-height: 1.8;">{getattr(project, 'purpose', '') if hasattr(project, 'purpose') else project.get('purpose', '未提供')}</p>
            </div>

            <div class="grid-2">
                <div class="stat-card">
                    <div class="stat-value" id="stat-experts">0</div>
                    <div class="stat-label">专家数量</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value" id="stat-criteria">0</div>
                    <div class="stat-label">准则数量</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value" id="stat-factors">0</div>
                    <div class="stat-label">因素数量</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value" id="stat-cr">-</div>
                    <div class="stat-label">一致性CR</div>
                </div>
            </div>

            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">关键因素 Top 5</h2>
                </div>
                <ul class="ranking-list" id="top5-list"></ul>
            </div>
        </div>

        <!-- Criteria Section -->
        <div id="criteria" class="section-content">
            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">准则层权重</h2>
                    <span class="badge badge-success" id="cr-badge">CR &le; 0.10</span>
                </div>
                <div id="criteria-bars"></div>
            </div>

            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">准则层权重详情</h2>
                </div>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>编号</th>
                            <th>准则名称</th>
                            <th>权重</th>
                            <th>λmax</th>
                            <th>CI</th>
                            <th>CR</th>
                        </tr>
                    </thead>
                    <tbody id="criteria-table-body"></tbody>
                </table>
            </div>
        </div>

        <!-- Alternatives Section -->
        <div id="alternatives" class="section-content">
            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">因素综合权重排名</h2>
                </div>
                <ul class="ranking-list" id="full-ranking-list"></ul>
            </div>

            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">权重分布</h2>
                </div>
                <div id="weight-distribution"></div>
            </div>
        </div>

        <!-- Convergence Section -->
        <div id="convergence" class="section-content">
            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">收敛检验结果</h2>
                </div>
                <div id="convergence-summary"></div>
            </div>

            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">各因素收敛状态</h2>
                </div>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>因素</th>
                            <th>均值</th>
                            <th>标准差</th>
                            <th>CV</th>
                            <th>一致性</th>
                            <th>状态</th>
                        </tr>
                    </thead>
                    <tbody id="convergence-table-body"></tbody>
                </table>
            </div>
        </div>

        <!-- Sensitivity Section -->
        <div id="sensitivity" class="section-content">
            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">敏感性分析结果</h2>
                </div>
                <p style="color: var(--text-secondary); margin-bottom: 1.5rem;">
                    分析方法：对准则权重和方案得分分别进行 &plusmn;10% 的变幅测试
                </p>
                <div id="sensitivity-summary"></div>
            </div>
        </div>

        <!-- Experts Section -->
        <div id="experts" class="section-content">
            <div class="card">
                <div class="card-header">
                    <h2 class="card-title">专家团队</h2>
                </div>
                <div id="expert-list"></div>
            </div>
        </div>
    </div>

    <div class="footer">
        <p>本报告由 Delphi-AHP 流程自动生成</p>
        <p>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>

    <script>
        // Report Data
        const reportData = {json_data};

        // Navigation
        function showSection(sectionId) {{
            document.querySelectorAll('.section-content').forEach(el => el.classList.remove('active'));
            document.querySelectorAll('.nav-tab').forEach(el => el.classList.remove('active'));
            document.getElementById(sectionId).classList.add('active');
            event.target.classList.add('active');
        }}

        // Initialize
        document.addEventListener('DOMContentLoaded', function() {{
            // Stats
            document.getElementById('stat-experts').textContent = reportData.experts.length;
            document.getElementById('stat-criteria').textContent = reportData.criteria.length;
            document.getElementById('stat-factors').textContent = reportData.alternatives.length;

            // CR value
            const ahpCons = reportData.convergence_data?.ahp_consistency;
            if (ahpCons) {{
                const crEl = document.getElementById('stat-cr');
                crEl.textContent = ahpCons.cr.toFixed(4);
                const badge = document.getElementById('cr-badge');
                if (!ahpCons.passed) {{
                    badge.className = 'badge badge-danger';
                    badge.textContent = 'CR > 0.10';
                }}
            }}

            // Top 5 Ranking
            const top5List = document.getElementById('top5-list');
            reportData.sorted_alts.slice(0, 5).forEach((alt, i) => {{
                const altWeight = reportData.ranking.find(r => r.id === alt.id);
                const rankClass = i === 0 ? 'rank-1' : i === 1 ? 'rank-2' : i === 2 ? 'rank-3' : 'rank-default';
                top5List.innerHTML += `
                    <li class="ranking-item">
                        <span class="rank-badge ${{rankClass}}">${{i + 1}}</span>
                        <div class="ranking-content">
                            <div class="ranking-name">${{alt.name}}</div>
                            <div class="ranking-meta">${{reportData.criteria_map?.[alt.belongs_to_criteria] || alt.belongs_to_criteria || ''}}</div>
                        </div>
                        <span class="ranking-score">${{(altWeight?.weight * 100 || 0).toFixed(2)}}%</span>
                    </li>
                `;
            }});

            // Criteria Bars
            const criteriaBars = document.getElementById('criteria-bars');
            reportData.criteria_weights?.forEach(c => {{
                criteriaBars.innerHTML += `
                    <div class="criteria-bar">
                        <div class="criteria-label">
                            <span>${{c.name}}</span>
                            <span>${{(c.weight * 100).toFixed(1)}}%</span>
                        </div>
                        <div class="progress-bar">
                            <div class="progress-fill" style="width: ${{c.weight * 100}}%"></div>
                        </div>
                    </div>
                `;
            }});

            // Criteria Table
            const criteriaTableBody = document.getElementById('criteria-table-body');
            reportData.criteria_weights?.forEach(c => {{
                criteriaTableBody.innerHTML += `
                    <tr>
                        <td>${{c.id}}</td>
                        <td>${{c.name}}</td>
                        <td>${{c.weight.toFixed(4)}}</td>
                        <td>${{(c.lambda_max || 0).toFixed(4)}}</td>
                        <td>${{(c.ci || 0).toFixed(4)}}</td>
                        <td>${{(c.cr || 0).toFixed(4)}} ${{c.consistency_passed ? '✓' : '✗'}}</td>
                    </tr>
                `;
            }});

            // Full Ranking List
            const fullRankingList = document.getElementById('full-ranking-list');
            reportData.sorted_alts.forEach((alt, i) => {{
                const altWeight = reportData.ranking.find(r => r.id === alt.id);
                const rankClass = i === 0 ? 'rank-1' : i === 1 ? 'rank-2' : i === 2 ? 'rank-3' : 'rank-default';
                fullRankingList.innerHTML += `
                    <li class="ranking-item">
                        <span class="rank-badge ${{rankClass}}">${{i + 1}}</span>
                        <div class="ranking-content">
                            <div class="ranking-name">${{alt.name}}</div>
                            <div class="ranking-meta">${{reportData.criteria_map?.[alt.belongs_to_criteria] || alt.belongs_to_criteria || ''}} | 原始得分: ${{(alt.raw_score || 0).toFixed(2)}}</div>
                        </div>
                        <span class="ranking-score">${{(altWeight?.weight * 100 || 0).toFixed(2)}}%</span>
                    </li>
                `;
            }});

            // Weight Distribution
            const weightDist = document.getElementById('weight-distribution');
            reportData.sorted_alts.slice(0, 10).forEach((alt, i) => {{
                const altWeight = reportData.ranking.find(r => r.id === alt.id);
                const weight = altWeight?.weight || 0;
                weightDist.innerHTML += `
                    <div class="criteria-bar">
                        <div class="criteria-label">
                            <span>${{i + 1}}. ${{alt.name}}</span>
                            <span>${{(weight * 100).toFixed(2)}}%</span>
                        </div>
                        <div class="progress-bar">
                            <div class="progress-fill" style="width: ${{weight * 100 * 3}}%"></div>
                        </div>
                    </div>
                `;
            }});

            // Convergence
            const convSummary = document.getElementById('convergence-summary');
            const convData = reportData.convergence_data;
            if (convData && convData.factor_results) {{
                const total = convData.total_factors || Object.keys(convData.factor_results).length;
                const converged = convData.converged_count || 0;
                convSummary.innerHTML = `
                    <div class="grid-2">
                        <div class="stat-card">
                            <div class="stat-value">${{converged}}/${{total}}</div>
                            <div class="stat-label">已收敛因素</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">${{(converged / total * 100).toFixed(0)}}%</div>
                            <div class="stat-label">收敛率</div>
                        </div>
                    </div>
                `;
            }}

            const convTableBody = document.getElementById('convergence-table-body');
            if (convData && convData.factor_results) {{
                Object.entries(convData.factor_results).forEach(([fid, result]) => {{
                    const statusClass = result.converged ? 'badge-success' : 'badge-danger';
                    const statusText = result.converged ? '已收敛' : '未收敛';
                    convTableBody.innerHTML += `
                        <tr>
                            <td>${{fid}}</td>
                            <td>${{(result.mean || 0).toFixed(2)}}</td>
                            <td>${{(result.std_dev || 0).toFixed(2)}}</td>
                            <td>${{(result.cv_percent || 0).toFixed(2)}}%</td>
                            <td>${{result.agreement || '-'}}</td>
                            <td><span class="badge ${{statusClass}}">${{statusText}}</span></td>
                        </tr>
                    `;
                }});
            }}

            // Sensitivity
            const sensSummary = document.getElementById('sensitivity-summary');
            if (reportData.sensitivity_data) {{
                const critSens = reportData.sensitivity_data.criteria_sensitivity;
                const scoreSens = reportData.sensitivity_data.score_sensitivity;

                let stableCount = 0;
                let totalCount = 0;

                critSens?.scenarios?.forEach(s => {{
                    totalCount += 2;
                    const upStable = Object.values(s.increase_scenario?.ranking_changes || {{}}).every(c => c.stable);
                    const downStable = Object.values(s.decrease_scenario?.ranking_changes || {{}}).every(c => c.stable);
                    if (upStable && downStable) stableCount++;
                }});

                scoreSens?.scenarios?.forEach(s => {{
                    totalCount += 2;
                    const upStable = Object.values(s.increase_scenario?.ranking_changes || {{}}).every(c => c.stable);
                    const downStable = Object.values(s.decrease_scenario?.ranking_changes || {{}}).every(c => c.stable);
                    if (upStable && downStable) stableCount++;
                }});

                const stablePct = totalCount > 0 ? (stableCount / totalCount * 100).toFixed(0) : 0;
                const stability = stablePct >= 80 ? '高' : stablePct >= 50 ? '中' : '低';

                sensSummary.innerHTML = `
                    <div class="grid-2">
                        <div class="stat-card">
                            <div class="stat-value">${{stableCount}}/${{totalCount}}</div>
                            <div class="stat-label">稳定场景</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">${{stablePct}}%</div>
                            <div class="stat-label">稳健性</div>
                        </div>
                    </div>
                    <p style="margin-top: 1rem; color: var(--text-secondary);">
                        稳健性${{stability}}：${{stablePct}}% 的场景在参数变化 &plusmn;10% 时排名保持稳定
                    </p>
                `;
            }}

            // Experts
            const expertList = document.getElementById('expert-list');
            reportData.experts?.forEach(exp => {{
                const initials = exp.name?.slice(0, 1) || '?';
                expertList.innerHTML += `
                    <div class="expert-card">
                        <div class="expert-avatar">${{initials}}</div>
                        <div class="expert-info">
                            <div class="expert-name">${{exp.name}}</div>
                            <div class="expert-role">${{exp.role}} | ${{exp.org_type}} | ${{exp.expertise}}</div>
                        </div>
                    </div>
                `;
            }});
        }});
    </script>
</body>
</html>
"""
    return html


def create_deliverable_zip(state: dict, run_dir: Path, csv_path: Path = None) -> Path:
    """
    Create a ZIP file containing all deliverables.
    """
    run_id = state.get("run_id", "unknown")
    zip_filename = f"Delphi_AHP_Report_{run_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    zip_path = run_dir / zip_filename

    # Find interview records (dialogue files)
    interview_files = []
    if (run_dir / "interview_records").exists():
        interview_files = list((run_dir / "interview_records").glob("*.md"))
    elif (run_dir / "interviews").exists():
        interview_files = list((run_dir / "interviews").glob("**/*.md"))

    # Also look for expert_dialogues directory
    expert_dialogues_dir = run_dir / EXPERT_DIALOGUES_DIR
    if expert_dialogues_dir.exists():
        interview_files.extend(list(expert_dialogues_dir.glob("*.md")))

    # Find interview framework/questionnaire
    framework_files = []
    for pattern in ["*framework*.json", "*protocol*.json", "*questionnaire*.json", "*访谈框架*", "*问题*"]:
        framework_files.extend(run_dir.glob(pattern))

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Add final report
        if (run_dir / "final_report.md").exists():
            zipf.write(run_dir / "final_report.md", "报告文档/final_report.md")

        # Add executive summary
        if (run_dir / "executive_summary.md").exists():
            zipf.write(run_dir / "executive_summary.md", "报告文档/executive_summary.md")

        # Add interactive HTML report
        if (run_dir / "interactive_report.html").exists():
            zipf.write(run_dir / "interactive_report.html", "报告文档/interactive_report.html")

        # Add analysis tables (step 5)
        if (run_dir / "analysis_results.md").exists():
            zipf.write(run_dir / "analysis_results.md", "分析表格/analysis_results.md")

        # Add other analysis files
        for artifact in ["convergence_check.json", "convergence_summary.md",
                         "sensitivity_analysis.json", "sensitivity_summary.md",
                         "ahp_results.json", "ahp_hierarchy.json",
                         "criteria_comparisons.json", "alternative_scores.json"]:
            if (run_dir / artifact).exists():
                if artifact.endswith('.json'):
                    zipf.write(run_dir / artifact, f"数据文件/{artifact}")
                else:
                    zipf.write(run_dir / artifact, f"分析表格/{artifact}")

        # Add CSV analysis tables
        if csv_path and csv_path.exists():
            zipf.write(csv_path, f"分析表格/{csv_path.name}")
        # Also add any other CSV files in the run_dir
        for csv_file in run_dir.glob("analysis_tables_*.csv"):
            zipf.write(csv_file, f"分析表格/{csv_file.name}")

        # Add XLSX analysis tables
        for xlsx_file in run_dir.glob("analysis_tables.xlsx"):
            zipf.write(xlsx_file, f"分析表格/{xlsx_file.name}")

        # Add interview framework/questionnaire
        for fw_file in framework_files:
            if fw_file.exists():
                zipf.write(fw_file, f"访谈框架/{fw_file.name}")

        # Add expert dialogue records
        for dialogue_file in interview_files:
            if dialogue_file.exists():
                # Preserve subdirectory structure if exists
                arcname = f"专家对话/{dialogue_file.name}"
                zipf.write(dialogue_file, arcname)

        # Add project configuration
        if (run_dir / "project.json").exists():
            zipf.write(run_dir / "project.json", "配置/project.json")

    return zip_path


def _read_csv_table(csv_path: Path) -> tuple:
    """Read a CSV file and return (headers, rows)."""
    import csv as csv_lib
    headers = []
    rows = []
    try:
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv_lib.reader(f)
            for i, row in enumerate(reader):
                if i == 0:
                    # Skip BOM row if present
                    if row and row[0] and row[0][0] == '\ufeff':
                        row[0] = row[0][1:]
                    headers = row
                else:
                    rows.append(row)
    except Exception as e:
        print(f"    [警告] 读取CSV失败 {csv_path.name}: {e}")
    return headers, rows


def _build_report_context(
    project_title: str,
    project: dict,
    hierarchy: dict,
    ahp_results: dict,
    criteria_comparisons: dict,
    alternative_scores: dict,
    convergence_data: dict,
    sensitivity_data: dict,
    expert_data: list,
    run_dir: Path,
) -> dict:
    """
    Build a comprehensive context dict with all data for LLM report generation.
    Reads CSV files from run_dir.
    """
    criteria = hierarchy.get("criteria_layer", [])
    alternatives = hierarchy.get("alternative_layer", [])
    criteria_weights_list = ahp_results.get("criteria", [])
    ranking = ahp_results.get("ranking", [])
    rank_map = {r["id"]: r["rank"] for r in ranking}
    sorted_alts = sorted(alternatives, key=lambda x: rank_map.get(x["id"], 999))

    # Read all CSV tables
    csv_tables = {}
    csv_files = list(run_dir.glob("analysis_tables_*.csv"))
    for csv_file in csv_files:
        headers, rows = _read_csv_table(csv_file)
        if headers:
            csv_tables[csv_file.name] = {"headers": headers, "rows": rows}

    # Build criteria comparison matrix text
    crit_ids = [c["id"] for c in criteria]
    n = len(criteria)
    crit_matrix_lines = ["## 准则层两两比较矩阵 ( Saaty 1-9 尺度)\n"]
    crit_matrix_lines.append("| 准则A | 准则B | Saaty值(几何平均) |")
    crit_matrix_lines.append("|:---|:---|:---:|")
    comparisons = criteria_comparisons.get("comparisons", {})
    expert_comparisons = criteria_comparisons.get("expert_comparisons", [])
    for i in range(n):
        for j in range(i + 1, n):
            key = f"{i},{j}"
            val = comparisons.get(key, 1.0)
            crit_matrix_lines.append(f"| {criteria[i]['name']} | {criteria[j]['name']} | {val:.4f} |")

    # Build criteria weights text
    crit_weights_lines = ["## 准则层权重\n"]
    crit_weights_lines.append("| 编号 | 准则名称 | 权重 | λmax | CI | CR | 一致性 |")
    crit_weights_lines.append("|:---:|:---|:---:|:---:|:---:|:---:|:---:|")
    crit_cons = ahp_results.get("criteria_consistency", {})
    for c in criteria:
        cw = next((cr for cr in criteria_weights_list if cr["id"] == c["id"]), {})
        cr_val = cw.get("cr", 0)
        passed = "✓ 通过" if cw.get("consistency_passed", False) else "✗ 未通过"
        crit_weights_lines.append(
            f"| {c['id']} | {c['name']} | {cw.get('weight', 0):.4f} | "
            f"{cw.get('lambda_max', 0):.4f} | {cw.get('ci', 0):.4f} | {cr_val:.4f} | {passed} |"
        )
    crit_weights_lines.append(
        f"| **汇总** | - | 1.0000 | {crit_cons.get('lambda_max', 0):.4f} | "
        f"{crit_cons.get('ci', 0):.4f} | **{crit_cons.get('cr', 0):.4f}** | "
        f"{'✓ 通过' if crit_cons.get('passed') else '✗ 未通过'} |"
    )

    # Build alternatives ranking text
    alt_ranking_lines = ["## 方案层综合权重排名\n"]
    alt_ranking_lines.append("| 排名 | 编号 | 方案名称 | 所属准则 | 综合权重 |")
    alt_ranking_lines.append("|:---:|:---:|:---|:---|:---:|")
    criteria_map = {c["id"]: c["name"] for c in criteria}
    for alt in sorted_alts:
        aw = next((a for a in ahp_results.get("alternatives", []) if a["id"] == alt["id"]), {})
        crit_name = criteria_map.get(alt.get("belongs_to_criteria", ""), "?")
        rank = rank_map.get(alt["id"], "?")
        combined = aw.get("combined_weight", 0)
        alt_ranking_lines.append(f"| {rank} | {alt['id']} | {alt['name']} | {crit_name} | {combined:.4f} |")

    # Convergence summary
    convergence_lines = []
    if convergence_data:
        total = convergence_data.get("total_factors", 0)
        converged = convergence_data.get("converged_count", 0)
        convergence_lines.append("## 收敛检验结果\n")
        convergence_lines.append(f"- 检验因素总数：{total}")
        convergence_lines.append(f"- 已收敛因素数：{converged} ({converged/total*100:.1f}%)" if total > 0 else "- 无收敛数据")
        if convergence_data.get("ahp_consistency"):
            ahp_c = convergence_data["ahp_consistency"]
            passed = "通过" if ahp_c.get("passed") else "未通过"
            convergence_lines.append(f"- AHP一致性检验：CR={ahp_c.get('cr', 0):.4f} ({passed})")

    # Sensitivity summary
    sensitivity_lines = []
    if sensitivity_data:
        sensitivity_lines.append("## 敏感性分析\n")
        sens_crit = sensitivity_data.get("criteria_sensitivity", [])
        if isinstance(sens_crit, list) and sens_crit:
            sensitivity_lines.append("### 准则层权重敏感性\n")
            sensitivity_lines.append("| 准则 | 基准权重 | 变动范围 | 最高权重 | 最低权重 |")
            sensitivity_lines.append("|:---|:---:|:---:|:---:|:---:|")
            for s in sens_crit[:5]:
                sensitivity_lines.append(
                    f"| {s.get('name','?')} | {s.get('base_weight',0):.4f} | "
                    f"±{s.get('variation',0):.2%} | {s.get('max_weight',0):.4f} | {s.get('min_weight',0):.4f} |"
                )

    return {
        "project_title": project_title,
        "background": project.get("background", "") if isinstance(project, dict) else getattr(project, "background", ""),
        "purpose": project.get("purpose", "") if isinstance(project, dict) else getattr(project, "purpose", ""),
        "criteria": criteria,
        "alternatives": alternatives,
        "criteria_weights": criteria_weights_list,
        "criteria_consistency": crit_cons,
        "ranking": ranking,
        "expert_data": expert_data,
        "n_experts": len(expert_data),
        "csv_tables": csv_tables,
        "criteria_comparison_text": "\n".join(crit_matrix_lines),
        "criteria_weights_text": "\n".join(crit_weights_lines),
        "alternatives_ranking_text": "\n".join(alt_ranking_lines),
        "convergence_text": "\n".join(convergence_lines) if convergence_lines else "",
        "sensitivity_text": "\n".join(sensitivity_lines) if sensitivity_lines else "",
    }


def generate_final_report_with_llm(
    state: dict,
    run_dir: Path,
    hierarchy: dict,
    ahp_results: dict,
    criteria_comparisons: dict,
    alternative_scores: dict,
    convergence_data: dict,
    sensitivity_data: dict,
    expert_data: list,
) -> str:
    """
    Generate final report using LLM, incorporating all analysis tables.
    """
    from llm import call_llm

    # Get provider from state or from providers.json file
    providers = state.get("providers", {})
    selected_provider = None

    if isinstance(providers, dict) and providers:
        # Find the selected/active provider from state
        for p in providers.values():
            if getattr(p, "key", None) == state.get("selected_provider_key"):
                selected_provider = p
                break
        if not selected_provider:
            # Try to find a configured one
            for p in providers.values():
                if getattr(p, "api_key", None):
                    selected_provider = p
                    break
    else:
        # Try loading from providers.json in run_dir
        providers_file = run_dir / "providers.json"
        if providers_file.exists():
            try:
                from providers import PROVIDER_TEMPLATES
                providers_data = load_json(providers_file)
                for key, prov_data in providers_data.items():
                    if prov_data.get("api_key") and prov_data.get("base_url"):
                        # Reconstruct Provider object
                        class _Prov:
                            pass
                        p = _Prov()
                        p.key = key
                        p.name = prov_data.get("name", key)
                        p.adapter = prov_data.get("adapter", "openai_chat_compatible")
                        p.base_url = prov_data.get("base_url", "")
                        p.api_key = prov_data.get("api_key", "")
                        p.default_model = prov_data.get("default_model", prov_data.get("model", ""))
                        p.models = prov_data.get("models", [])
                        selected_provider = p
                        break
            except Exception as e:
                print(f"    [警告] 加载providers.json失败: {e}")

    if selected_provider:
        base_url = getattr(selected_provider, "base_url", "")
        api_key = getattr(selected_provider, "api_key", "")
        model = getattr(selected_provider, "default_model", "")
    else:
        print("    [警告] 未找到可用的LLM提供商，跳过LLM报告生成")
        return None

    project = state.get("project", {})
    project_title = project.get("title", "未知项目") if isinstance(project, dict) else getattr(project, "title", "未知项目")

    # Build context
    ctx = _build_report_context(
        project_title, project, hierarchy, ahp_results,
        criteria_comparisons, alternative_scores, convergence_data,
        sensitivity_data, expert_data, run_dir,
    )

    # Build CSV tables overview for prompt
    csv_overview_lines = ["### 已生成的CSV分析表格：\n"]
    for fname, data in ctx["csv_tables"].items():
        csv_overview_lines.append(f"**{fname}** ({len(data['rows'])} 行数据)：")
        if data["headers"]:
            csv_overview_lines.append(f"  列：{', '.join(data['headers'][:6])}{'...' if len(data['headers']) > 6 else ''}")
    csv_overview = "\n".join(csv_overview_lines)

    system_prompt = """你是一位专业的德尔菲-层次分析法（Delphi-AHP）研究报告撰写专家。
你的任务是根据提供的研究数据，撰写一份完整、深入、专业的Markdown格式研究报告。
报告应该：
1. 结构清晰，层次分明
2. 包含详细的数据解读和分析
3. 对关键发现提供解释和洞察
4. 使用规范的学术语言
5. 在适当的地方使用表格、列表等格式化元素
6. 生成的内容应直接用于最终报告，不需要进一步修改"""

    user_prompt = f"""## 研究项目信息
- **项目标题**：{ctx['project_title']}
- **研究背景**：{ctx['background'] or '未提供'}
- **研究目的**：{ctx['purpose'] or '未提供'}
- **专家人数**：{ctx['n_experts']}位
- **准则数量**：{len(ctx['criteria'])}个
- **方案数量**：{len(ctx['alternatives'])}个

## 准则层信息
"""
    for c in ctx["criteria"]:
        user_prompt += f"- **{c['id']}** {c['name']}: {c.get('description', '')[:50]}\n"

    user_prompt += f"""
## 专家团队
"""
    for e in ctx["expert_data"]:
        user_prompt += f"- {e['name']}（{e['role']}，{e['org_type']}，专长：{e['expertise']}）\n"

    user_prompt += f"""
## 准则层两两比较矩阵
{ctx['criteria_comparison_text']}

## 准则层权重及一致性
{ctx['criteria_weights_text']}

## 方案层综合权重排名
{ctx['alternatives_ranking_text']}
"""

    if ctx["convergence_text"]:
        user_prompt += f"\n{ctx['convergence_text']}\n"

    if ctx["sensitivity_text"]:
        user_prompt += f"\n{ctx['sensitivity_text']}\n"

    user_prompt += f"""
## 已生成的分析表格文件
{csv_overview}

---
请撰写一份完整的Delphi-AHP研究报告，包含以下章节：
1. 研究概述（背景、目的、方法）
2. 专家团队介绍
3. 德尔菲法收敛结果
4. AHP分析结果（准则层权重、方案层排名）
5. 一致性检验分析
6. 综合讨论与结论

报告语言为中文，使用规范的学术写作风格。
所有表格数据必须正确呈现。
"""

    print(f"    正在调用LLM生成报告（模型：{model}）...")
    try:
        report_content = call_llm(
            base_url=base_url,
            api_key=api_key,
            model=model,
            prompt=user_prompt,
            system_prompt=system_prompt,
            temperature=0.7,
            timeout=180,
        )
        return report_content
    except Exception as e:
        print(f"    [警告] LLM报告生成失败: {e}")
        return None


def run_step8(state: dict) -> dict:
    """
    Run step 8: Generate final report and deliverables.

    This step automatically:
    1. Loads all data from previous steps
    2. Generates comprehensive markdown report
    3. Generates executive summary
    4. Generates interactive HTML report
    5. Creates ZIP archive with all deliverables
    6. Copies to current directory

    Args:
        state: Current state dict

    Returns:
        Updated state dict
    """
    print_step_header()

    run_id = state.get("run_id", "unknown")
    run_dir = Path(state.get("run_dir", Path(__file__).parent.parent / "run_result" / run_id))

    print(f"Run ID: {run_id}")
    print(f"运行目录: {run_dir}")
    print()

    if not run_dir.exists():
        print(f"  [错误] 找不到运行目录: {run_dir}")
        return state

    # Generate main report
    print("=" * 60)
    print("  【生成报告】")
    print("=" * 60)
    print()

    project = state.get("project", {})
    project_title = project.get('title', '未知项目') if isinstance(project, dict) else getattr(project, 'title', '未知项目')

    # Load all data for report generation
    hierarchy = {}
    ahp_results = {}
    criteria_comparisons = {}
    alternative_scores = {}
    convergence_data = {}
    sensitivity_data = {}
    expert_data = []

    if (run_dir / "ahp_hierarchy.json").exists():
        hierarchy = load_json(run_dir / "ahp_hierarchy.json")

    if (run_dir / "ahp_results.json").exists():
        ahp_results = load_json(run_dir / "ahp_results.json")

    if (run_dir / "criteria_comparisons.json").exists():
        criteria_comparisons = load_json(run_dir / "criteria_comparisons.json")

    if (run_dir / "alternative_scores.json").exists():
        alternative_scores = load_json(run_dir / "alternative_scores.json") or {}

    if (run_dir / "convergence_check.json").exists():
        convergence_data = load_json(run_dir / "convergence_check.json")

    if (run_dir / "sensitivity_analysis.json").exists():
        sensitivity_data = load_json(run_dir / "sensitivity_analysis.json")

    # Get experts info
    experts = state.get("experts", [])
    for exp in experts:
        expert_dict = {
            "id": getattr(exp, 'id', '?'),
            "name": getattr(exp, 'name', '未知'),
            "role": getattr(exp, 'role', '未知'),
            "org_type": getattr(exp, 'org_type', '未知'),
            "expertise": getattr(exp, 'expertise', '未知'),
        }
        expert_data.append(expert_dict)

    # Build criteria map
    criteria = hierarchy.get("criteria_layer", [])
    alternatives = hierarchy.get("alternative_layer", [])
    criteria_map = {c["id"]: c["name"] for c in criteria}

    # Ranking
    ranking = ahp_results.get("ranking", [])
    rank_map = {r["id"]: r["rank"] for r in ranking}
    sorted_alts = sorted(alternatives, key=lambda x: rank_map.get(x["id"], 999))

    # Generate report content using LLM
    print("  正在通过LLM生成完整研究报告...")
    criteria_weights = ahp_results.get("criteria", [])
    report_content = generate_final_report_with_llm(
        state=state,
        run_dir=run_dir,
        hierarchy=hierarchy,
        ahp_results=ahp_results,
        criteria_comparisons=criteria_comparisons,
        alternative_scores=alternative_scores,
        convergence_data=convergence_data,
        sensitivity_data=sensitivity_data,
        expert_data=expert_data,
    )

    # Fallback if LLM failed
    if not report_content:
        print("    [提示] LLM生成失败，使用简化模板")
        report_lines = []
        report_lines.append(f"# Delphi-AHP 研究报告\n")
        report_lines.append(f"**项目标题**: {project_title}")
        report_lines.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        report_lines.append("---\n")
        report_lines.append("## 1. 研究概述\n")
        report_lines.append(f"**研究背景**: {project.get('background', '未提供')}\n")
        report_lines.append(f"**研究目的**: {project.get('purpose', '未提供')}\n")
        report_lines.append("---\n")
        report_lines.append("## 2. 专家团队\n")
        report_lines.append(f"本研究邀请了 {len(expert_data)} 位专家参与评估。\n")
        report_lines.append("## 3. AHP 分析结果\n")
        report_lines.append("### 准则层权重\n")
        for c in criteria:
            cw = next((cr for cr in criteria_weights if cr["id"] == c["id"]), {})
            report_lines.append(f"- {c['name']}: {cw.get('weight', 0):.4f}")
        report_lines.append("\n### 方案层权重排名\n")
        for alt in sorted_alts:
            aw = next((a for a in ahp_results.get("alternatives", []) if a["id"] == alt["id"]), {})
            rank = rank_map.get(alt["id"], "?")
            report_lines.append(f"{rank}. {alt['name']}: {aw.get('combined_weight', 0):.4f}")
        report_lines.append(f"\n---\n*本报告由Delphi-AHP流程自动生成*")
        report_content = "\n".join(report_lines)

    # Save markdown report
    report_path = run_dir / "final_report.md"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_content)
    print(f"  [已生成] {report_path}")

    # Generate executive summary
    summary_lines = []
    summary_lines.append(f"# 执行摘要：{project_title}\n")
    summary_lines.append("## 关键因素 Top 5\n")
    for i, alt in enumerate(sorted_alts[:5], 1):
        aw = next((a for a in ahp_results.get("alternatives", []) if a["id"] == alt["id"]), {})
        combined = aw.get("combined_weight", 0)
        summary_lines.append(f"{i}. {alt['name']}（权重：{combined:.4f}，{combined*100:.1f}%）")

    summary_lines.append("\n## 准则层权重\n")
    for c in criteria:
        cw = next((cr for cr in criteria_weights if cr["id"] == c["id"]), {})
        weight = cw.get("weight", 0)
        summary_lines.append(f"- {c['name']}：{weight:.2%}")

    summary_lines.append("\n---\n*详细分析见完整报告*")

    summary_content = "\n".join(summary_lines)
    summary_path = run_dir / "executive_summary.md"
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write(summary_content)
    print(f"  [已生成] {summary_path}")

    # Generate interactive HTML report
    print("  正在生成交互式HTML报告...")
    html_content = generate_interactive_html_report(state, run_dir)
    html_path = run_dir / "interactive_report.html"
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"  [已生成] {html_path}")

    # Generate CSV analysis tables
    print("  正在生成CSV分析表格...")
    csv_path = generate_analysis_csv(state, run_dir)
    if csv_path:
        print(f"  [已生成] {csv_path}")

    # XLSX is generated inside generate_analysis_csv
    xlsx_path = run_dir / "analysis_tables.xlsx"

    # Create expert_dialogues directory and save expert dialogues
    print("  正在保存专家对话记录...")
    dialogues_dir = run_dir / EXPERT_DIALOGUES_DIR
    dialogues_dir.mkdir(parents=True, exist_ok=True)

    # Try to save expert dialogues from interview records
    interview_records = load_json(run_dir / "interview_records.json") if (run_dir / "interview_records.json").exists() else {}
    records = interview_records.get("records", [])
    for record in records:
        expert_name = record.get("expert_name", "unknown")
        expert_id = record.get("expert_id", "E00")
        qa_pairs = record.get("qa_pairs", [])

        # Create a markdown file for each expert's dialogue
        dialogue_lines = []
        dialogue_lines.append(f"# 专家对话记录：{expert_name}\n")
        dialogue_lines.append(f"**专家ID**: {expert_id}\n")
        dialogue_lines.append("---\n")

        for i, qa in enumerate(qa_pairs, 1):
            round_info = qa.get("round", "第一轮")
            question = qa.get("question", "未提供问题")
            answer = qa.get("answer", "未提供回答")

            dialogue_lines.append(f"## {round_info} - Q{i}\n")
            dialogue_lines.append(f"**问题**: {question}\n")
            dialogue_lines.append(f"\n**回答**: {answer}\n")
            dialogue_lines.append("---\n")

        dialogue_path = dialogues_dir / f"dialogue_{expert_id}_{expert_name}.md"
        with open(dialogue_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(dialogue_lines))
        print(f"  [已保存] {dialogue_path}")

    # Create ZIP archive
    print("\n  正在创建交付压缩包...")
    zip_path = create_deliverable_zip(state, run_dir, csv_path)
    print(f"  [已生成] {zip_path}")

    # Copy to run directory
    dest_dir = run_dir
    dest_report = dest_dir / f"report_{run_id}.md"
    dest_summary = dest_dir / f"summary_{run_id}.md"
    dest_html = dest_dir / f"interactive_report_{run_id}.html"
    dest_xlsx = dest_dir / "analysis_tables.xlsx"
    dest_zip = dest_dir / zip_path.name

    try:
        shutil.copy(report_path, dest_report)
        print(f"\n  [已复制] {dest_report}")
        shutil.copy(summary_path, dest_summary)
        print(f"  [已复制] {dest_summary}")
        shutil.copy(html_path, dest_html)
        print(f"  [已复制] {dest_html}")
        if xlsx_path.exists():
            shutil.copy(xlsx_path, dest_xlsx)
            print(f"  [已复制] {dest_xlsx}")
        shutil.copy(zip_path, dest_zip)
        print(f"  [已复制] {dest_zip}")
    except Exception as e:
        print(f"\n  [提示] 部分文件无法复制到当前目录: {e}")

    # Update summary
    summary_data = {
        "run_id": run_id,
        "project_title": project_title,
        "status": "completed",
        "created_at": datetime.now().isoformat(),
        "artifacts": [
            "final_report.md",
            "executive_summary.md",
            "interactive_report.html",
            "analysis_tables.csv",
            "analysis_tables.xlsx",
            EXPERT_DIALOGUES_DIR,
            zip_path.name,
        ]
    }
    save_json(run_dir / "report_summary.json", summary_data)

    # Display preview
    print("\n" + "=" * 60)
    print("  【报告预览】")
    print("=" * 60)
    print()

    for line in report_content.split('\n')[:40]:
        print(line)

    if len(report_content.split('\n')) > 40:
        print(f"\n... (共 {len(report_content.split(chr(10)))} 行)")

    state["run_status"] = "completed"
    state["report_path"] = str(report_path)

    print()
    print("=" * 60)
    print("  步骤 8 完成！")
    print("=" * 60)
    print(f"\n  已生成文件：")
    print(f"    - final_report.md (完整报告)")
    print(f"    - executive_summary.md (执行摘要)")
    print(f"    - interactive_report.html (交互式HTML报告)")
    print(f"    - analysis_tables.csv (分析表格CSV)")
    print(f"    - analysis_tables.xlsx (分析表格XLSX)")
    print(f"    - {EXPERT_DIALOGUES_DIR}/ (专家对话记录目录)")
    print(f"    - {zip_path.name} (交付压缩包)")
    print(f"\n  压缩包包含：")
    print(f"    - 报告文档（markdown格式）")
    print(f"    - 交互式HTML报告")
    print(f"    - 分析表格（CSV + XLSX格式）")
    print(f"    - 数据文件（JSON格式）")
    print(f"    - 访谈框架（问题设计）")
    print(f"    - 专家对话记录")

    return state
